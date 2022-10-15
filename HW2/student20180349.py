#!/usr/bin/python3
import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

row_id = 1;


for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id,column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
	row_id += 1

for row in ws:
	if sum_v >= 90:
		grade = "A"
	elif sum_v >= 80:
		grade = "B"
	elif sum_v >= 70:
		grade ="C"
	else:
		grade ="D"
	ws.cell(row=row_id, column = 8).value = grade

wb.save("student.xlsx")
